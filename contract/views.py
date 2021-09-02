# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.http import HttpResponse
from django.template import RequestContext
from django.db.models.query_utils import Q
from django.contrib.auth.decorators import login_required
from contract.models import table, detail
from vacation.models import user_table
from wzhl_oa.settings import BASE_DIR
import json
import datetime
import os
from libs.common import Num2MoneyFormat
from django import forms
from openpyxl import load_workbook
from threading import Thread
from libs.sendmail import send_mail


@login_required
def contract_apply(request):
    path = request.path.split('/')[1]
    return render(request, 'contract/contract_table.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'contract',
                                                 'path2':path,
                                                 'page_name1':u'合同管理',
                                                 'page_name2':u'合同申请',},
                                                context_instance=RequestContext(request))




@login_required
def contract_apply_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['contract_uuid','apply_time','name','contract_class','party_b','contract_name','contract_amount_figures','status','id']

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name)
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    else:
        if sSearch == '':
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name)
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]


    for i in  result_data:
        aaData.append({
                       '0':i.contract_uuid,
                       '1':str(i.apply_time),
                       '2':i.name,
                       '3':i.contract_class,
                       '4':i.party_b,
                       '5':i.contract_name,
                       '6':i.contract_amount_figures,
                       '7':i.status,
                       '8':i.id,
                       '9':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")




@login_required
def contract_set_session(request):
    _id = request.POST.get('id')
    commit = request.POST.get('commit')
    if _id == '0':
        try:
            request.session.pop('contract_id')
        except KeyError:
            pass
    elif _id:
        request.session['contract_id'] = int(_id)
    if commit:
        request.session['contract_commit'] = commit
    return HttpResponse(json.dumps('OK'),content_type="application/json")



@login_required
def contract_apply_detail(request):
    path = request.path.split('/')[1]
    contract_uuid = request.POST.get('contract_uuid')
    origin_contract_uuid = request.POST.get('origin_contract_uuid')
    party_a = request.POST.get('party_a')
    # name = request.POST.get('name')
    # department = request.POST.get('department')
    finance_class = request.POST.get('finance_class')
    contract_class = request.POST.get('contract_class')
    contract_name = request.POST.get('contract_name')
    party_b = request.POST.get('party_b')
    address = request.POST.get('address')
    contacts = request.POST.get('contacts')
    e_mail = request.POST.get('e_mail')
    phone_1 = request.POST.get('phone_1')
    phone_2 = request.POST.get('phone_2')
    fax = request.POST.get('fax')
    bank = request.POST.get('bank')
    bank_account = request.POST.get('bank_account')
    contract_detail = request.POST.get('contract_detail')
    currency_type = request.POST.get('currency_type')
    contract_amount_figures = request.POST.get('contract_amount_figures')
    special_requirements = request.POST.get('special_requirements')
    contract_begin_time = request.POST.get('begin')
    contract_end_time = request.POST.get('end')
    partner_qualification = request.POST.get('partner_qualification')
    comment = request.POST.get('comment')
    _id = request.POST.get('id')
    commit = request.POST.get('commit')

    user_info_orm = user_table.objects.get(name=request.user.first_name)

    name = user_info_orm.name
    department = user_info_orm.department
    apply_time = ''
    contract_amount_words = ''
    stamp_status = 0
    archive_status = 0
    status = 0
    if contract_amount_figures:
        if float(contract_amount_figures) == 0.0:
            finance_class = u'无金额'

    try:
        request.session.pop('contract_result')
    except KeyError:
        pass

    table_id =  request.session.get('contract_id')
    if not table_id:
        if party_a and finance_class and contract_class and contract_name and party_b and contacts and phone_1 and contract_detail and contract_amount_figures:
            try:
                if party_a == u'上海六界信息技术有限公司':
                    uuid_a = 'SHLJ'
                elif party_a == u'北京七葫芦科技有限公司':
                    uuid_a = 'BJZF'
                elif party_a == u'霍尔果斯柒色葫芦广告科技有限公司':
                    uuid_a = 'HCQSHL'
                elif party_a == u'Six World Inc.':
                    uuid_a = 'SWI'
                elif party_a == u'六界娱乐有限公司（SIX WORLD ENTERTAINMENT LIMITED）':
                    uuid_a = 'LJYL'
                elif party_a == u'上海陆色网络科技有限公司':
                    uuid_a = 'SHLS'
                elif party_a == u'上海果豆网络科技有限公司':
                    uuid_a = 'SHGD'
                elif party_a == u'北京葫芦豆科技有限公司':
                    uuid_a = 'BJHLD'
                elif party_a == u'上海六赞网络科技有限公司':
                    uuid_a = 'SHLZ'
                elif party_a == u'北京六娱信息技术有限公司':
                    uuid_a = 'BJLY'
                elif party_a == u'浙江星果文化传媒有限公司':
                    uuid_a = 'ZJXG'
                elif party_a == u'浙江六趣网络科技有限公司':
                    uuid_a = 'ZJLQ'

                uuid_b = {u'商务合作':'SWHZ',
                          u'艺人经纪':'YRJJ',
                          u'市场框架':'SCKJ',
                          u'技术开发':'JSKF',
                          u'行政':'XZ',
                          u'人事':'RS',
                          u'财务':'CW',
                          u'法务':'FW',
                          u'其他':'QT'}

                if len(table.objects.filter(contract_uuid__contains=uuid_b[contract_class])) == 0:
                    uuid_c = str(datetime.datetime.now().year) + '0001'
                else:
                    uuid_c_orm = table.objects.filter(contract_uuid__contains=uuid_b[contract_class]).order_by('id').reverse()[0]
                    uuid_c = str(int(str(datetime.datetime.now().year) + uuid_c_orm.contract_uuid.split('-')[2][4:]) + 1)

                contract_uuid = '-'.join([uuid_a, uuid_b[contract_class], uuid_c])
                print contract_uuid

                contract_amount_words = Num2MoneyFormat(float(contract_amount_figures))

                if currency_type != '美元':
                    if party_a in [u'上海六界信息技术有限公司', u'北京七葫芦科技有限公司', 'Six World Inc.', u'六界娱乐有限公司（SIX WORLD ENTERTAINMENT LIMITED）',
                                   u'上海陆色网络科技有限公司',u'上海果豆网络科技有限公司',u'北京葫芦豆科技有限公司',u'上海六赞网络科技有限公司',
                                   u'北京六娱信息技术有限公司',u'浙江星果文化传媒有限公司',u'浙江六趣网络科技有限公司']:
                        if finance_class == u'无金额':
                            process_type = 'l'
                        else:
                            if float(contract_amount_figures) >= 10000:
                                process_type = 'l'
                            else:
                                process_type = 's'
                    elif party_a == u'霍尔果斯柒色葫芦广告科技有限公司':
                        if finance_class == u'无金额':
                            process_type = 'l'
                        if finance_class == u'收':
                            if float(contract_amount_figures) >= 500000:
                                process_type = 'l'
                            else:
                                process_type = 's'
                        if finance_class == u'支':
                            if float(contract_amount_figures) >= 200000:
                                process_type = 'l'
                            else:
                                process_type = 's'
                else:
                    process_type = 'l'
                if party_a == u'北京七葫芦科技有限公司':
                    status = 1
                    approve_now = u'卞蓓'
                elif (user_info_orm.supervisor == u'曹津' and process_type == 'l') or user_info_orm.supervisor == user_info_orm.principal:
                    print user_info_orm.supervisor
                    status = 1
                    approve_now = user_info_orm.principal
                else:
                    status = -15
                    approve_now = user_info_orm.supervisor

                orm = table(party_a=party_a,name=name,department=department,finance_class=finance_class,contract_class=contract_class,\
                            contract_uuid=contract_uuid,origin_contract_uuid=origin_contract_uuid,contract_name=contract_name,\
                            party_b=party_b,address=address,contacts=contacts,e_mail=e_mail,phone_1=phone_1,phone_2=phone_2,\
                            fax=fax,bank=bank,bank_account=bank_account,comment=comment,contract_detail=contract_detail,\
                            currency_type=currency_type,contract_amount_figures=contract_amount_figures,contract_amount_words=contract_amount_words,\
                            special_requirements=special_requirements,contract_begin_time=contract_begin_time,\
                            contract_end_time=contract_end_time,partner_qualification=partner_qualification,stamp_status=0,\
                            archive_status=0,status=status,approve_now=approve_now,commit_time=datetime.datetime.now(),process_type=process_type)

                orm.save()

                try:
                    archive_path = request.session['contract_upload_file']
                    request.session['contract_upload_file'] = ''
                except KeyError:
                    archive_path = ''

                orm_last_id = table.objects.all().order_by('id').reverse()[0]

                orm2 = detail(name=request.user.first_name,operation=9,archive_path=archive_path,comment='',parent_id=orm_last_id.id)
                orm2.save()

                request.session['contract_result'] = u'保存成功'
            except Exception,e:
                print e
                request.session['contract_result'] = e
        else:
            if commit == '1':
                request.session['contract_result'] = u'星号为必填项'
    else:
        try:
            orm = table.objects.get(id=table_id)
            commit = request.session.get('contract_commit')
            print commit

            if commit != '0':
                orm.party_a = party_a
                orm.finance_class = finance_class
                orm.contract_class = contract_class
                orm.origin_contract_uuid = origin_contract_uuid
                orm.contract_name = contract_name
                orm.party_b = party_b
                orm.address = address
                orm.contacts = contacts
                orm.e_mail = e_mail
                orm.phone_1 = phone_1
                orm.phone_2 = phone_2
                orm.fax = fax
                orm.bank = bank
                orm.bank_account = bank_account
                orm.comment = comment
                orm.contract_detail = contract_detail
                orm.currency_type = currency_type
                orm.contract_amount_figures = contract_amount_figures
                contract_amount_words = Num2MoneyFormat(float(contract_amount_figures))
                orm.contract_amount_words = contract_amount_words
                orm.special_requirements = special_requirements
                orm.contract_begin_time = contract_begin_time
                orm.contract_end_time = contract_end_time
                orm.partner_qualification = partner_qualification
                orm.apply_time = datetime.datetime.now()
                orm.comment = comment

                if currency_type != '美元':
                    if party_a in [u'上海六界信息技术有限公司', 'Six World Inc.', u'六界娱乐有限公司（SIX WORLD ENTERTAINMENT LIMITED）',
                                   u'上海陆色网络科技有限公司',u'上海果豆网络科技有限公司',u'北京葫芦豆科技有限公司',u'上海六赞网络科技有限公司',
                                   u'北京六娱信息技术有限公司',u'浙江星果文化传媒有限公司',u'浙江六趣网络科技有限公司']:
                        if finance_class == u'无金额':
                            process_type = 'l'
                        else:
                            if float(contract_amount_figures) >= 10000:
                                process_type = 'l'
                            else:
                                process_type = 's'
                    elif party_a == u'北京七葫芦科技有限公司' or party_a == u'霍尔果斯柒色葫芦广告科技有限公司':
                        if finance_class == u'无金额':
                            process_type = 'l'
                        if finance_class == u'收':
                            if float(contract_amount_figures) >= 500000:
                                process_type = 'l'
                            else:
                                process_type = 's'
                        if finance_class == u'支':
                            if float(contract_amount_figures) >= 200000:
                                process_type = 'l'
                            else:
                                process_type = 's'
                else:
                    process_type = 'l'

                orm.process_type = process_type

                # if orm.status == 0:
                #     if user_info_orm.principal == u'曹津' and process_type == 'l':
                #         print user_info_orm.principal
                #         orm.status = 2
                #         orm.approve_now = u'龚晓芸'
                #     else:
                #         orm.status = 1
                #         orm.approve_now = user_info_orm.principal

                orm.save()
                # request.session['contract_result'] = u'保存成功'
            # try:
            #     request.session.pop('contract_commit')
            # except KeyError:
            #     pass
            return render(request, 'contract/contract_table_detail.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                     'path1':'contract',
                                                     'path2':path,
                                                     'page_name1':u'合同管理',
                                                     'page_name2':u'合同信息',
                                                     'name':orm.name,
                                                     'department':orm.department,
                                                     'apply_time':str(orm.apply_time),
                                                     'party_a':orm.party_a,
                                                     'finance_class':orm.finance_class,
                                                     'contract_uuid':orm.contract_uuid,
                                                     'origin_contract_uuid':orm.origin_contract_uuid,
                                                     'contract_class':orm.contract_class,
                                                     'party_b':orm.party_b,
                                                     'contacts':orm.contacts,
                                                     'contract_name':orm.contract_name,
                                                     'address':orm.address,
                                                     'e_mail':orm.e_mail,
                                                     'phone_1':orm.phone_1,
                                                     'phone_2':orm.phone_2,
                                                     'fax':orm.fax,
                                                     'bank':orm.bank,
                                                     'bank_account':orm.bank_account,
                                                     'comment':orm.comment.replace('\r\n','\\n'),
                                                     'contract_detail':orm.contract_detail.replace('\r\n','\\n'),
                                                     'currency_type':orm.currency_type,
                                                     'contract_amount_figures':orm.contract_amount_figures,
                                                     'contract_amount_words':orm.contract_amount_words,
                                                     'special_requirements':orm.special_requirements.replace('\r\n','\\n'),
                                                     'contract_begin_time':str(orm.contract_begin_time),
                                                     'contract_end_time':str(orm.contract_end_time),
                                                     'partner_qualification':orm.partner_qualification,
                                                     'stamp_status':orm.stamp_status,
                                                     'archive_status':orm.archive_status,
                                                     'status':orm.status,
                                                     'process_type':orm.process_type,
                                                     'id':orm.id
                                                     },
                                                    context_instance=RequestContext(request))
        except Exception,e:
            print e
            request.session['contract_result'] = e
    return render(request, 'contract/contract_table_detail.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                     'path1':'contract',
                                                     'path2':path,
                                                     'page_name1':u'合同管理',
                                                     'page_name2':u'合同信息',
                                                     'name':name,
                                                     'department':department,
                                                     'apply_time':apply_time,
                                                     'contract_amount_words':contract_amount_words,
                                                     'stamp_status':stamp_status,
                                                     'archive_status':archive_status,
                                                     'status':status,
                                                     },
                                                    context_instance=RequestContext(request))





@login_required
def contract_approve(request):
    path = request.path.split('/')[1]
    return render(request, 'contract/contract_approve.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'contract',
                                                 'path2':path,
                                                 'page_name1':u'合同管理',
                                                 'page_name2':u'全部合同',},
                                                context_instance=RequestContext(request))




@login_required
def contract_approve_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['contract_uuid','apply_time','name','contract_class','party_b','contract_name','contract_amount_figures','status','id']

    subordinate = []
    subordinate_orm = user_table.objects.filter(principal=request.user.first_name)
    for i in subordinate_orm:
        subordinate.append(i.name)

    # if request.user.has_perm('contract.can_view_all'):
    #     if  sSortDir_0 == 'asc':
    #         if sSearch == '':
    #             result_data = table.objects.all().order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
    #             iTotalRecords = table.objects.all().count()
    #         else:
    #             result_data = table.objects.all()
    #             sSearch_list = sSearch.split()
    #             for i in range(len(sSearch_list)):
    #                 result_data = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
    #                                                 Q(party_b__contains=sSearch_list[i]) | \
    #                                                 Q(contract_name__contains=sSearch_list[i]))
    #
    #                 iTotalRecords = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
    #                                                 Q(party_b__contains=sSearch_list[i]) | \
    #                                                 Q(contract_name__contains=sSearch_list[i])).count()
    #             result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    #     else:
    #         if sSearch == '':
    #             result_data = table.objects.all().order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    #             iTotalRecords = table.objects.all().count()
    #         else:
    #             result_data = table.objects.all()
    #             sSearch_list = sSearch.split()
    #             for i in range(len(sSearch_list)):
    #                 result_data = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
    #                                                 Q(party_b__contains=sSearch_list[i]) | \
    #                                                 Q(contract_name__contains=sSearch_list[i]))
    #
    #                 iTotalRecords = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
    #                                                 Q(party_b__contains=sSearch_list[i]) | \
    #                                                 Q(contract_name__contains=sSearch_list[i])).count()
    #             result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    # else:
    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = table.objects.filter(approve_now=request.user.first_name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(approve_now=request.user.first_name).count()
        else:
            result_data = table.objects.all()
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(approve_now=request.user.first_name).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(approve_now=request.user.first_name).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    else:
        if sSearch == '':
            result_data = table.objects.filter(approve_now=request.user.first_name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(approve_now=request.user.first_name).count()
        else:
            result_data = table.objects.all()
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(approve_now=request.user.first_name).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(approve_now=request.user.first_name).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                Q(party_b__contains=sSearch_list[i]) | \
                                                Q(contract_name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]


    for i in  result_data:
        export = '''
                    <a class="btn btn-sm green">
                        生成Excel文件 <i class="fa fa-level-down"></i>
                    </a>
                '''
        approve = '''
                    <a class="btn btn-sm blue">
                        审批 <i class="fa"></i>
                    </a>
                 '''
        aaData.append({
                       '0':i.contract_uuid,
                       '1':str(i.apply_time),
                       '2':i.name,
                       '3':i.contract_class,
                       '4':i.party_b,
                       '5':i.contract_name,
                       '6':i.contract_amount_figures,
                       '7':i.status,
                       '8':export,
                       '9':approve,
                       '10':i.id,
                       '11':i.process_type,
                       '12':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")




@login_required
def contract_all(request):
    path = request.path.split('/')[1]
    return render(request, 'contract/contract_all.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'contract',
                                                 'path2':path,
                                                 'page_name1':u'合同管理',
                                                 'page_name2':u'全部合同',},
                                                context_instance=RequestContext(request))




@login_required
def contract_all_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['contract_uuid','apply_time','name','contract_class','party_b','contract_name','contract_amount_figures','status','id']

    subordinate = []
    subordinate_orm = user_table.objects.filter(principal=request.user.first_name)
    for i in subordinate_orm:
        subordinate.append(i.name)

    if request.user.has_perm('contract.can_view_all'):
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.all().order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
        else:
            if sSearch == '':
                result_data = table.objects.all().order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.all().filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    elif request.user.has_perm('contract.can_view_part'):
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
                                                   Q(contract_uuid__contains='HCQSHL')).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
                                                   Q(contract_uuid__contains='HCQSHL')).count()
            else:
                result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
                                                   Q(contract_uuid__contains='HCQSHL'))
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
        else:
            if sSearch == '':
                result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
                                                   Q(contract_uuid__contains='HCQSHL')).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
                                                   Q(contract_uuid__contains='HCQSHL')).count()
            else:
                result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
                                                   Q(contract_uuid__contains='HCQSHL'))
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    elif request.user.has_perm('contract.can_view_liujie'):
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.filter(contract_uuid__contains='SHLJ').order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(contract_uuid__contains='SHLJ').count()
            else:
                result_data = table.objects.filter(contract_uuid__contains='SHLJ')
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
        else:
            if sSearch == '':
                result_data = table.objects.filter(contract_uuid__contains='SHLJ').order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(contract_uuid__contains='SHLJ').count()
            else:
                result_data = table.objects.filter(contract_uuid__contains='SHLJ')
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    elif request.user.has_perm('contract.can_view_shangwu'):
        orm = user_table.objects.filter(department=u'商务中心')
        shangwu_name_list = []
        for i in orm:
            shangwu_name_list.append(i.name)
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.filter(name__in=shangwu_name_list).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(name__in=shangwu_name_list).count()
            else:
                result_data = table.objects.filter(name__in=shangwu_name_list)
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
        else:
            if sSearch == '':
                result_data = table.objects.filter(name__in=shangwu_name_list).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(name__in=shangwu_name_list).count()
            else:
                result_data = table.objects.filter(name__in=shangwu_name_list)
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    else:
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.filter(name__in=subordinate).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(name__in=subordinate).count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(name__in=subordinate).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(name__in=subordinate).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
        else:
            if sSearch == '':
                result_data = table.objects.filter(name__in=subordinate).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(name__in=subordinate).count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(name__in=subordinate).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(name__in=subordinate).filter(Q(contract_uuid__contains=sSearch_list[i]) | \
                                                    Q(party_b__contains=sSearch_list[i]) | \
                                                    Q(contract_name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]


    for i in  result_data:
        export = '''
                    <a class="btn btn-sm green">
                        生成Excel文件 <i class="fa fa-level-down"></i>
                    </a>
                '''
        approve = '''
                    <a class="btn btn-sm blue">
                        审批 <i class="fa"></i>
                    </a>
                 '''
        aaData.append({
                       '0':i.contract_uuid,
                       '1':str(i.apply_time),
                       '2':i.name,
                       '3':i.contract_class,
                       '4':i.party_b,
                       '5':i.contract_name,
                       '6':i.contract_amount_figures,
                       '7':i.status,
                       '8':export,
                       '9':approve,
                       '10':i.id,
                       '11':i.process_type,
                       '12':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")



class UploadFileForm(forms.Form):
    title = forms.CharField(max_length=50)
    file = forms.FileField()

@login_required
def handle_uploaded_file(request,f):
    file_name = ''
    try:
        path = 'media/'
        # file_name = path + f.name
        today = datetime.datetime.now()
        print 'mkdir -p {0}/{1}{2}/{3}/{4}/'.format(BASE_DIR, path, today.year, today.month, today.day)
        os.system('mkdir -p {0}/{1}{2}/{3}/{4}/'.format(BASE_DIR, path, today.year, today.month, today.day))
        full_name = '{0}{1}/{2}/{3}/{4}'.format(path, today.year, today.month, today.day, f.name)
        time = datetime.datetime.now().strftime('%H%M%S')
        if os.path.isfile(BASE_DIR + '/' + full_name):
            full_name =  '{0}{1}/{2}/{3}/{4}_{5}'.format(path, today.year, today.month, today.day, time, f.name)
            # orm = upload_files.objects.get(file_name=f.name)
            # orm.file_name = f.name + '_' + time
            # orm.save()
        file = open(BASE_DIR + '/' + full_name, 'wb+')
        for chunk in f.chunks():
            file.write(chunk)
        file.close()
        # file_size = os.path.getsize(BASE_DIR + full_name)
        # upload_files.objects.create(file_name=f.name,file_size=file_size,upload_user=request.user.username)

        request.session['contract_upload_file'] = full_name
        result_code = 0
    except Exception, e:
        import traceback
        print traceback.format_exc()
        # logger.error(e)
        result_code = 1
    return result_code

@login_required
def contract_get_upload(request):
    file = request.FILES.get('file')
    if not file == None:
        result_code = handle_uploaded_file(request,file)
        if result_code == 0:
            return HttpResponse(json.dumps({'msg': "上传成功", "code": 0}),content_type="application/json")
        else:
            return HttpResponse(json.dumps({'msg': "上传失败", "code": 1}),content_type="application/json")
    else:
        return HttpResponse(json.dumps({'msg': "上传失败", "code": 1}),content_type="application/json")



@login_required
def contract_approve_process(request):
    flag = request.POST.get('flag')
    _id = request.POST.get('id')
    status = request.POST.get('status')
    process_type = request.POST.get('process_type')
    comment = request.POST.get('comment')
    add_process = request.POST.get('add_process')

    print _id,type(_id)


    try:
        try:
            archive_path = request.session['contract_upload_file']
            request.session['contract_upload_file'] = ''
        except KeyError:
            archive_path = ''

        orm = table.objects.get(id=_id)

        if orm.party_a in [u'上海六界信息技术有限公司', 'Six World Inc.', u'六界娱乐有限公司（SIX WORLD ENTERTAINMENT LIMITED）',
                           u'上海陆色网络科技有限公司',u'上海果豆网络科技有限公司',u'北京葫芦豆科技有限公司',u'上海六赞网络科技有限公司',
                           u'北京六娱信息技术有限公司',u'浙江星果文化传媒有限公司',u'浙江六趣网络科技有限公司']:
            status_owner = {2: u'龚晓芸',
                            3: u'高茹',
                            4: u'吴佳伟',
                            6: u'王海峰',
                            7: u'曹津',
                            8: u'吴佳伟',
                            9: u'吴佳伟'}
        elif orm.party_a == u'北京七葫芦科技有限公司':
            status_owner = {2: u'龚晓芸',
                            3: u'高茹',
                            4: u'吴佳伟',
                            6: u'王海峰',
                            7: u'曹津',
                            8: u'卞蓓',
                            9: u'吴佳伟'}
        else:
            status_owner = {2: u'龚晓芸',
                            3: u'高茹',
                            4: u'吴佳伟',
                            6: u'王海峰',
                            7: u'曹津',
                            8: u'卞蓓',
                            9: u'卞蓓'}

        print status
        if request.user.first_name != orm.approve_now:
            if not (status == '10' and orm.name == request.user.first_name):
                return HttpResponse(json.dumps({'code':1,'msg':u'您不是审批人'}),content_type="application/json")

        detail_orm = detail.objects.filter(parent_id=_id)
        if flag == '1':
            if status == '-1':
                principal_orm = user_table.objects.get(name=orm.name)
                print principal_orm.principal
                if orm.currency_type != '美元':
                    if orm.party_a in [u'上海六界信息技术有限公司', 'Six World Inc.', u'六界娱乐有限公司（SIX WORLD ENTERTAINMENT LIMITED）',
                                       u'上海陆色网络科技有限公司',u'上海果豆网络科技有限公司',u'北京葫芦豆科技有限公司',
                                       u'上海六赞网络科技有限公司',u'北京六娱信息技术有限公司',u'浙江星果文化传媒有限公司',u'浙江六趣网络科技有限公司']:
                        if orm.finance_class == u'无金额':
                            orm.process_type = 'l'
                        else:
                            if float(orm.contract_amount_figures) >= 10000:
                                orm.process_type = 'l'
                            else:
                                orm.process_type = 's'
                    elif orm.party_a == u'北京七葫芦科技有限公司' or orm.party_a == u'霍尔果斯柒色葫芦广告科技有限公司':
                        if orm.finance_class == u'无金额':
                            orm.process_type = 'l'
                        if orm.finance_class == u'收':
                            if float(orm.contract_amount_figures) >= 500000:
                                orm.process_type = 'l'
                            else:
                                orm.process_type = 's'
                        if orm.finance_class == u'支':
                            if float(orm.contract_amount_figures) >= 200000:
                                orm.process_type = 'l'
                            else:
                                orm.process_type = 's'
                else:
                    orm.process_type = 'l'
                orm.save()

                if orm.party_a == u'北京七葫芦科技有限公司':
                    orm.status = 1
                    orm.approve_now = u'卞蓓'
                elif (principal_orm.supervisor == u'曹津' and process_type == 'l') or principal_orm.supervisor == principal_orm.principal:
                    orm.status = 1
                    orm.approve_now = principal_orm.principal
                else:
                    orm.status = -15
                    orm.approve_now = principal_orm.supervisor

                flag = '-1'

            if status == '-15':
                principal_orm = user_table.objects.get(name=orm.name)
                orm.status = 1
                orm.approve_now = principal_orm.principal

            if status == '1':
                # for i in detail_orm:
                #     if status_owner[2] == i.name:
                #         if process_type == 'l':
                #             orm.status = 3
                #         else:
                #             orm.status = 4
                #         break
                # else:
                #     orm.status = 2
                orm.status = 2
                orm.approve_now = status_owner[orm.status]

            if status == '2':
                # for i in detail_orm:
                #     if process_type == 'l':
                #         if status_owner[3] == i.name:
                #             orm.status = 4
                #             break
                #         else:
                #              orm.status = 3
                #     else:
                #         if status_owner[4] == i.name:
                #             orm.status = 6
                #             break
                #         else:
                #             orm.status = 4
                if process_type == 'l':
                    orm.status = 3
                else:
                    orm.status = 4
                orm.approve_now = status_owner[orm.status]

            if status == '3':
                # for i in detail_orm:
                #     if status_owner[4] == i.name:
                #         orm.status = 6
                #         break
                # else:
                #     orm.status = 4
                orm.status = 4
                orm.approve_now = status_owner[orm.status]

            if status == '4':
                if add_process:
                    orm.status = 5
                    orm.approve_now = add_process
                else:
                    # if process_type == 'l':
                    #     for i in detail_orm:
                    #         if status_owner[6] == i.name:
                    #             orm.status = 7
                    #             break
                    #     else:
                    #         orm.status = 6
                    # else:
                    #     orm.status = 8
                    if process_type == 'l':
                        orm.status = 6
                    else:
                        orm.status = 8
                    orm.approve_now = status_owner[orm.status]

            if status == '5':
                # if process_type == 'l':
                #     for i in detail_orm:
                #         if status_owner[6] == i.name:
                #             orm.status = 7
                #             break
                #     else:
                #         orm.status = 6
                # else:
                #     orm.status = 8
                if process_type == 'l':
                    orm.status = 6
                else:
                    orm.status = 8
                orm.approve_now = status_owner[orm.status]

            if status == '6':
                # for i in detail_orm:
                #     if status_owner[7] == i.name:
                #         orm.status = 8
                #         break
                # else:
                #     orm.status = 7
                orm.status = 7
                orm.approve_now = status_owner[orm.status]

            if status == '7':
                orm.status = 8
                orm.approve_now = status_owner[orm.status]

            if status == '8':
                orm.status = 80
                orm.approve_now = orm.name
                orm.stamp_status = 1

            if status == '80':
                orm.status = 9
                orm.approve_now = status_owner[orm.status]
                orm.stamp_status = 1

            if status == '9':
                orm.status = 10
                orm.approve_now = ''
                orm.archive_status = 1

            if status == '31':
                orm.status = 32
                if orm.name != status_owner[4]:
                    orm.approve_now = status_owner[4]
                else:
                    orm.approve_now = status_owner[6]

            if status == '32':
                orm.status = 33
                orm.approve_now = ''

        if flag == '2':
            if int(status) <= 4:
                orm.status = -1
                orm.approve_now = orm.name
            else:
                orm.status = 4
                orm.approve_now = status_owner[orm.status]

        if flag == '3':
                orm.status = 31
                if orm.name != status_owner[2]:
                    orm.approve_now = status_owner[2]
                else:
                    orm.approve_now = status_owner[3]

        if flag == '0':
            orm.status = 11
            orm.approve_now = ''

        if orm.approve_now:
            approve_now_orm = user_table.objects.get(name=orm.approve_now)
            email = approve_now_orm.email
            Thread(target=send_mail, args=(email, '合同审批提醒',
                                       '<h3>有一个合同事件等待您的审批，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/contract_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()

        orm.apply_time = datetime.datetime.now()
        orm.save()

        orm2 = detail(name=request.user.first_name,operation=flag,archive_path=archive_path,comment=comment,parent_id=_id)
        orm2.save()

        return HttpResponse(json.dumps({'code':0,'msg':u'审批成功'}),content_type="application/json")
    except Exception,e:
        import traceback
        print traceback.format_exc()
        print e
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")



@login_required
def contract_process_detail_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['apply_time','name','operation','archive_path','comment','id']

    try:
        parent_id = request.session['contract_id']
    except KeyError:
        parent_id = None
    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = detail.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).count()
        else:
            result_data = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(archive_path__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)) \
                                            .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(archive_path__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = detail.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).count()
        else:
            result_data = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(archive_path__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)) \
                                            .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(archive_path__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)).count()

    for i in  result_data:
        aaData.append({
                       '0':str(i.apply_time),
                       '1':i.name,
                       '2':i.operation,
                       '3':i.archive_path,
                       '4':i.comment.replace('\n','</br>'),
                       '5':i.id
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")





@login_required
def contract_create_excel(request):
    _id = request.POST.get('id')
    orm = table.objects.get(id=_id)

    if orm.status != 10:
        return HttpResponse(json.dumps({'code':1,'msg':'您的流程未完成无法生成Excel'}),content_type="application/json")

    contract_uuid = orm.contract_uuid
    origin_contract_uuid = orm.origin_contract_uuid
    party_a = orm.party_a
    apply_time = str(orm.apply_time)
    name = orm.name
    department = orm.department
    finance_class = orm.finance_class
    contract_class = orm.contract_class
    contract_name = orm.contract_name
    party_b = orm.party_b
    address = orm.address
    contacts = orm.contacts
    e_mail = orm.e_mail
    phone_1 = orm.phone_1
    phone_2 = orm.phone_2
    fax = orm.fax
    bank = orm.bank
    bank_account = orm.bank_account
    comment = orm.comment
    contract_detail = orm.contract_detail
    contract_amount_figures = orm.contract_amount_figures
    contract_amount_words = orm.contract_amount_words
    special_requirements = orm.special_requirements
    contract_begin_time = str(orm.contract_begin_time)
    contract_end_time = str(orm.contract_end_time)
    partner_qualification = orm.partner_qualification

    orm2 = detail.objects.filter(parent_id=_id)
    process_detail_list = []
    for i in orm2:
        process_detail_list.append({'apply_time':i.apply_time,
                                    'name':i.name,
                                    'operation':i.operation,
                                    'archive_path':i.archive_path,
                                    'comment':i.comment})

    try:
        wb = load_workbook(filename = BASE_DIR + '/static/files/contract_template.xlsx')
        ws = wb.active
        ws['B2'] = contract_uuid
        ws['F2'] = origin_contract_uuid
        ws['B3'] = party_a
        ws['F3'] = apply_time
        ws['B4'] = name
        ws['F4'] = department
        ws['B5'] = finance_class
        ws['F5'] = contract_class
        ws['B6'] = contract_name
        ws['C7'] = party_b
        ws['C8'] = address
        ws['C9'] = contacts
        ws['F9'] = e_mail
        ws['C10'] = phone_1
        ws['F10'] = phone_2
        ws['C11'] = fax
        ws['C12'] = bank
        ws['F12'] = bank_account
        ws['B13'] = comment
        ws['C14'] = contract_detail
        ws['D15'] = contract_amount_figures
        ws['F15'] = contract_amount_words
        ws['C16'] = special_requirements
        ws['C17'] = contract_begin_time
        ws['F17'] = contract_end_time
        ws['B18'] = partner_qualification

        row_num = 23
        for i in process_detail_list:
            ws['B'+str(row_num)] = i['apply_time']
            ws['C'+str(row_num)] = i['name']
            if i['operation'] == 1:
                operation = u'审批通过'
            elif i['operation'] == 0:
                operation = u'审批不通过'
            elif i['operation'] == -1:
                operation = u'退回修改'
            elif i['operation'] == 9:
                operation = u'合同申请'
            ws['D'+str(row_num)] = operation
            archive_path = os.path.basename(i['archive_path'])
            ws['E'+str(row_num)] = archive_path
            ws['F'+str(row_num)] = i['comment']
            row_num += 1

        wb.save(BASE_DIR + '/static/files/contract.xlsx')
        return HttpResponse(json.dumps({'code':0,'msg':u'生成成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")



@login_required
def contract_approve_alert(request):
    orm = table.objects.filter(approve_now=request.user.first_name)
    if len(orm) > 0:
        msg = '有%s个合同事件等待您的审批' % len(orm)
        return HttpResponse(json.dumps({'code':0,'msg':msg}),content_type="application/json")
    else:
        return HttpResponse(json.dumps({'code':1}),content_type="application/json")



@login_required
def export_contract_list(request):
    year = request.POST.get('year')
    contract_class = request.POST.get('contract_class')
    flag = ''
    if year:
        flag += 'x'
    if contract_class:
        flag += 'y'
    if flag:
        if flag == 'x':
            orm = table.objects.filter(apply_time__year=int(year)).filter(status=10)
        if flag == 'y':
            orm = table.objects.filter(contract_class=contract_class).filter(status=10)
        if flag == 'xy':
            orm = table.objects.filter(apply_time__year=int(year)).filter(contract_class=contract_class).filter(status=10)
    else:
        orm = table.objects.filter(status=10)
    contract_list = []
    for i in orm:
        contract_list.append({'contract_uuid':i.contract_uuid,
                              'apply_time':str(i.apply_time),
                              'name':i.name,
                              'contract_class':i.contract_class,
                              'party_b':i.party_b,
                              'contract_name':i.contract_name,
                              'contract_amount_figures':i.contract_amount_figures})
    try:
        wb = load_workbook(filename = BASE_DIR + '/static/files/contract_list_template.xlsx')
        ws = wb.active

        row_num = 3
        for i in contract_list:
            ws['A'+str(row_num)] = i['contract_uuid']
            ws['B'+str(row_num)] = i['apply_time']
            ws['C'+str(row_num)] = i['name']
            ws['D'+str(row_num)] = i['contract_class']
            ws['E'+str(row_num)] = i['party_b']
            ws['F'+str(row_num)] = i['contract_name']
            ws['G'+str(row_num)] = i['contract_amount_figures']
            ws['H'+str(row_num)] = '已归档'
            row_num += 1

        wb.save(BASE_DIR + '/static/files/contract_list.xlsx')
        return HttpResponse(json.dumps({'code':0,'msg':u'生成成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")