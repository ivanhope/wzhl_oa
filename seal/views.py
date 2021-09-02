# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.http import HttpResponse
from django.template import RequestContext
from django.db.models.query_utils import Q
from django.contrib.auth.decorators import login_required
from seal.models import table, detail
from vacation.models import user_table
from wzhl_oa.settings import BASE_DIR
import json
import datetime
import os
from libs.common import Num2MoneyFormat
from django import forms
from openpyxl import load_workbook
from threading import Thread
from libs.sendmail import send_mail


@login_required
def seal_apply(request):
    path = request.path.split('/')[1]
    return render(request, 'seal/seal_table.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'seal',
                                                 'path2':path,
                                                 'page_name1':u'印章管理',
                                                 'page_name2':u'印章申请',},
                                                context_instance=RequestContext(request))



@login_required
def seal_apply_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['uuid','name','department','seal_class','usage','status','id']

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name)
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    else:
        if sSearch == '':
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name)
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]


    for i in  result_data:
        aaData.append({
                       '0':i.uuid,
                       '1':i.name,
                       '2':i.reason,
                       '3':i.seal_class,
                       '4':i.usage,
                       '5':i.status,
                       '6':i.id,
                       '7':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")




@login_required
def seal_set_session(request):
    _id = request.POST.get('id')
    commit = request.POST.get('commit')
    if _id == '0':
        try:
            request.session.pop('seal_id')
        except KeyError:
            pass
    elif _id:
        request.session['seal_id'] = int(_id)
    if commit:
        request.session['seal_commit'] = commit
    return HttpResponse(json.dumps('OK'),content_type="application/json")




@login_required
def seal_apply_detail(request):
    path = request.path.split('/')[1]
    # uuid = request.POST.get('uuid')
    # name = request.POST.get('name')
    # department = request.POST.get('department')
    seal_from = request.POST.get('seal_from')
    seal_class = request.POST.get('seal_class')
    usage = request.POST.get('usage')
    reason = request.POST.get('reason')
    borrow_begin_time = request.POST.get('begin')
    borrow_end_time = request.POST.get('end')
    comment = request.POST.get('comment')
    _id = request.POST.get('id')
    commit = request.POST.get('commit')

    user_info_orm = user_table.objects.get(name=request.user.first_name)

    name = user_info_orm.name
    department = user_info_orm.department

    status = 0

    if usage == u'使用':
        borrow_begin_time = datetime.date(1970,1,1)
        borrow_end_time = datetime.date(1970,1,1)

    try:
        request.session.pop('seal_result')
    except KeyError:
        pass
    table_id =  request.session.get('seal_id')
    if not table_id:
        if seal_from and seal_class and usage and reason and borrow_begin_time and borrow_end_time:
            try:
                year = datetime.datetime.now().year
                if len(table.objects.filter(uuid__contains=year)) == 0:
                    uuid = str(datetime.datetime.now().year) + '-0001'
                else:
                    uuid_orm = table.objects.filter(uuid__contains=year).order_by('id').reverse()[0]
                    num = str(int(uuid_orm.uuid.split('-')[1]) + 1)
                    if len(num) < 4:
                        num = '0' * (4 - len(num)) + num
                    uuid = str(datetime.datetime.now().year) + '-' + num


                status = 1
                approve_now = user_info_orm.supervisor
                email_orm = user_table.objects.get(name=user_info_orm.supervisor)
                Thread(target=send_mail,args=(email_orm.email,'印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()

                try:
                    archive_path = request.session['seal_upload_file']
                    request.session['seal_upload_file'] = ''
                except KeyError:
                    archive_path = ''

                orm = table(uuid=uuid,name=name,department=department,seal_from=seal_from,seal_class=seal_class,\
                            usage=usage,reason=reason,archive_path=archive_path,borrow_begin_time=borrow_begin_time,\
                            borrow_end_time=borrow_end_time,comment=comment,status=status,approve_now=approve_now)

                orm.save()


                orm_last_id = table.objects.all().order_by('id').reverse()[0]

                orm2 = detail(name=request.user.first_name,operation=9,comment='',parent_id=orm_last_id.id)
                orm2.save()

                request.session['seal_result'] = u'保存成功'
            except Exception,e:
                print e
                request.session['seal_result'] = e
        else:
            if commit == '1':
                request.session['seal_result'] = u'不能留空'
    else:
        try:
            orm = table.objects.get(id=table_id)
            if seal_from and seal_class and usage and reason:
                commit = request.session.get('seal_commit')

                if commit != '0':
                    try:
                        archive_path = request.session['seal_upload_file']
                        request.session['seal_upload_file'] = ''
                    except KeyError:
                        if orm.archive_path:
                            archive_path = orm.archive_path
                        else:
                            archive_path = ''

                    orm.seal_from = seal_from
                    orm.seal_class = seal_class
                    orm.usage = usage
                    orm.reason = reason
                    orm.archive_path = archive_path
                    if borrow_begin_time and borrow_end_time:
                        orm.borrow_begin_time = borrow_begin_time
                        orm.borrow_end_time = borrow_end_time
                    orm.comment = comment

                    # if orm.status == 0:
                    #     if user_info_orm.principal == u'曹津' and process_type == 'l':
                    #         print user_info_orm.principal
                    #         orm.status = 2
                    #         orm.approve_now = u'龚晓芸'
                    #     else:
                    #         orm.status = 1
                    #         orm.approve_now = user_info_orm.principal

                    orm.save()
                    request.session['contract_result'] = u'保存成功'
            else:
                if commit == '1':
                    request.session['seal_result'] = u'不能留空'
            # try:
            #     request.session.pop('contract_commit')
            # except KeyError:
            #     pass
            return render(request, 'seal/seal_table_detail.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                     'path1':'seal',
                                                     'path2':path,
                                                     'page_name1':u'印章管理',
                                                     'page_name2':u'印章信息',
                                                     'name':orm.name,
                                                     'department':orm.department,
                                                     'uuid':orm.uuid,
                                                     'seal_from':orm.seal_from,
                                                     'seal_class':orm.seal_class,
                                                     'usage':orm.usage,
                                                                  # orm.reason
                                                     'reason':orm.reason,
                                                     # 'reason':u''+(json.dumps({'value':orm.reason},ensure_ascii=False)),
                                                     'archive_path':orm.archive_path,
                                                     'borrow_begin_time':str(orm.borrow_begin_time),
                                                     'borrow_end_time':str(orm.borrow_end_time),
                                                     'comment':orm.comment,
                                                     'status':orm.status,
                                                     'id':orm.id
                                                     },
                                                    context_instance=RequestContext(request))
        except Exception,e:
            print e
            request.session['contract_result'] = e
    return render(request, 'seal/seal_table_detail.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                     'path1':'seal',
                                                     'path2':path,
                                                     'page_name1':u'印章管理',
                                                     'page_name2':u'印章信息',
                                                     'name':name,
                                                     'department':department,
                                                     'status':status,
                                                     },
                                                    context_instance=RequestContext(request))




@login_required
def seal_apply_detail_sub(request):
    path = request.path.split('/')[1]
    # try:
    #     orm = user_table.objects.get(name=request.user.first_name)
    # except Exception:
    #     return render(request,'public/no_passing.html')
    table_id =  request.session.get('seal_id')
    if not table_id:
        try:
            archive_path = request.session['seal_upload_file']
        except KeyError:
            archive_path = ''
    else:
        orm = table.objects.get(id=table_id)
        archive_path = orm.archive_path
    return render(request, 'seal/seal_table_detail_sub.html',{'archive_path':archive_path,
                                                              'archive_path_filename':os.path.basename(archive_path)},context_instance=RequestContext(request))




@login_required
def seal_approve(request):
    path = request.path.split('/')[1]
    return render(request, 'seal/seal_approve.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'seal',
                                                 'path2':path,
                                                 'page_name1':u'印章管理',
                                                 'page_name2':u'全部印章',},
                                                context_instance=RequestContext(request))




@login_required
def seal_approve_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['uuid','name','department','seal_class','usage','status',None,None,'id']

    subordinate = []
    subordinate_orm = user_table.objects.filter(principal=request.user.first_name)
    for i in subordinate_orm:
        subordinate.append(i.name)

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = table.objects.filter(approve_now=request.user.first_name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(approve_now=request.user.first_name).count()
        else:
            result_data = table.objects.all()
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(approve_now=request.user.first_name).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(approve_now=request.user.first_name).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    else:
        if sSearch == '':
            result_data = table.objects.filter(approve_now=request.user.first_name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(approve_now=request.user.first_name).count()
        else:
            result_data = table.objects.all()
            sSearch_list = sSearch.split()
            for i in range(len(sSearch_list)):
                result_data = result_data.filter(approve_now=request.user.first_name).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                iTotalRecords = result_data.filter(approve_now=request.user.first_name).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
            result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]


    for i in  result_data:
        export = '''
                    <a class="btn btn-sm green">
                        生成Excel文件 <i class="fa fa-level-down"></i>
                    </a>
                '''
        approve = '''
                    <a class="btn btn-sm blue">
                        审批 <i class="fa"></i>
                    </a>
                 '''
        aaData.append({
                       '0':i.uuid,
                       '1':i.name,
                       '2':i.reason,
                       '3':i.seal_class,
                       '4':i.usage,
                       '5':i.status,
                       '6':export,
                       '7':approve,
                       '8':i.id,
                       '9':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")




@login_required
def seal_all(request):
    path = request.path.split('/')[1]
    return render(request, 'seal/seal_all.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'seal',
                                                 'path2':path,
                                                 'page_name1':u'印章管理',
                                                 'page_name2':u'全部印章',},
                                                context_instance=RequestContext(request))




@login_required
def seal_all_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['uuid','name','department','seal_class','usage','status',None,'id']

    subordinate = []
    subordinate_orm = user_table.objects.filter(principal=request.user.first_name)
    for i in subordinate_orm:
        subordinate.append(i.name)

    if request.user.has_perm('seal.can_view_all'):
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.all().order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.all().filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.all().filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
        else:
            if sSearch == '':
                result_data = table.objects.all().order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.all().filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.all().filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    # elif request.user.has_perm('seal.can_view_part'):
    #     if  sSortDir_0 == 'asc':
    #         if sSearch == '':
    #             result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
    #                                                Q(contract_uuid__contains='HCQSHL')).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
    #             iTotalRecords = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
    #                                                Q(contract_uuid__contains='HCQSHL')).count()
    #         else:
    #             result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
    #                                                Q(contract_uuid__contains='HCQSHL'))
    #             sSearch_list = sSearch.split()
    #             for i in range(len(sSearch_list)):
    #                 result_data = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
    #                                             Q(name__contains=sSearch_list[i]))
    #
    #                 iTotalRecords = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
    #                                             Q(name__contains=sSearch_list[i])).count()
    #             result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    #     else:
    #         if sSearch == '':
    #             result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
    #                                                Q(contract_uuid__contains='HCQSHL')).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    #             iTotalRecords = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
    #                                                Q(contract_uuid__contains='HCQSHL')).count()
    #         else:
    #             result_data = table.objects.filter(Q(contract_uuid__contains='BJZF') | \
    #                                                Q(contract_uuid__contains='HCQSHL'))
    #             sSearch_list = sSearch.split()
    #             for i in range(len(sSearch_list)):
    #                 result_data = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
    #                                             Q(name__contains=sSearch_list[i]))
    #
    #                 iTotalRecords = result_data.filter(Q(uuid__contains=sSearch_list[i]) | \
    #                                             Q(name__contains=sSearch_list[i])).count()
    #             result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
    else:
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                result_data = table.objects.filter(name__in=subordinate).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(name__in=subordinate).count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(name__in=subordinate).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(name__in=subordinate).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
        else:
            if sSearch == '':
                result_data = table.objects.filter(name__in=subordinate).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(name__in=subordinate).count()
            else:
                result_data = table.objects.all()
                sSearch_list = sSearch.split()
                for i in range(len(sSearch_list)):
                    result_data = result_data.filter(name__in=subordinate).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i]))

                    iTotalRecords = result_data.filter(name__in=subordinate).filter(Q(uuid__contains=sSearch_list[i]) | \
                                                Q(name__contains=sSearch_list[i])).count()
                result_data = result_data.order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]


    for i in  result_data:
        export = '''
                    <a class="btn btn-sm green">
                        生成Excel文件 <i class="fa fa-level-down"></i>
                    </a>
                '''
        aaData.append({
                       '0':i.uuid,
                       '1':i.name,
                       '2':i.reason,
                       '3':i.seal_class,
                       '4':i.usage,
                       '5':i.status,
                       '6':export,
                       '7':i.id,
                       '8':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")




class UploadFileForm(forms.Form):
    title = forms.CharField(max_length=50)
    file = forms.FileField()

@login_required
def handle_uploaded_file(request,f):
    file_name = ''
    try:
        path = 'media/'
        # file_name = path + f.name
        today = datetime.datetime.now()
        os.system('mkdir -p {0}/{1}{2}/{3}/{4}/'.format(BASE_DIR, path, today.year, today.month, today.day))
        full_name = '{0}/{1}{2}/{3}/{4}/{5}'.format(BASE_DIR, path, today.year, today.month, today.day, f.name)
        if os.path.isfile(full_name):
            time = datetime.datetime.now().strftime('%H%M%S')
            full_name =  '{0}/{1}{2}/{3}/{4}/{5}_{6}'.format(BASE_DIR, path, today.year, today.month, today.day, time, f.name)
            # orm = upload_files.objects.get(file_name=f.name)
            # orm.file_name = f.name + '_' + time
            # orm.save()
        file = open(full_name, 'wb+')
        for chunk in f.chunks():
            file.write(chunk)
        file.close()
        file_size = os.path.getsize(full_name)
        # upload_files.objects.create(file_name=f.name,file_size=file_size,upload_user=request.user.username)

        request.session['seal_upload_file'] = '{0}{1}/{2}/{3}/{4}'.format(path, today.year, today.month, today.day, f.name)
        result_code = 0
    except Exception, e:
        import traceback
        print traceback.format_exc()
        # logger.error(e)
        result_code = 1
    return result_code



@login_required
def seal_get_upload(request):
    file = request.FILES.get('file')
    if not file == None:
        result_code = handle_uploaded_file(request,file)
        if result_code == 0:
            return HttpResponse(json.dumps({'msg': "上传成功", "code": 0}),content_type="application/json")
        else:
            return HttpResponse(json.dumps({'msg': "上传失败", "code": 1}),content_type="application/json")
    else:
        return HttpResponse(json.dumps({'msg': "上传失败", "code": 1}),content_type="application/json")





@login_required
def seal_approve_process(request):
    flag = request.POST.get('flag')
    _id = request.POST.get('id')
    status = request.POST.get('status')
    comment = request.POST.get('comment')
    add_process = request.POST.get('add_process')

    print _id,type(_id)


    try:
        orm = table.objects.get(id=_id)

        if request.user.first_name != orm.approve_now:
            return HttpResponse(json.dumps({'code':1,'msg':u'您不是审批人'}),content_type="application/json")

        detail_orm = detail.objects.filter(parent_id=_id)
        if flag == '1':
            if status == '-1':
                supervisor_orm = user_table.objects.get(name=orm.name)
                print supervisor_orm.supervisor

                orm.status = 1
                orm.approve_now = supervisor_orm.supervisor
                email_orm = user_table.objects.get(name=supervisor_orm.supervisor)
                flag = '-1'
                Thread(target=send_mail,args=(email_orm.email,'印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()

            if status == '1':
                if orm.seal_class == u'法人章':
                    if orm.name == u'龚晓芸':
                        orm.status = 5
                        orm.approve_now = ''
                    else:
                        orm.status = 2
                        orm.approve_now = u'龚晓芸'
                        Thread(target=send_mail,args=('gongxiaoyun@xiaohulu.com','印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                else:
                    if orm.name == u'吴佳伟':
                        orm.status = 5
                        orm.approve_now = ''
                    else:
                        orm.status = 2
                        orm.approve_now = u'吴佳伟'
                        Thread(target=send_mail,args=('zhangliying@xiaohulu.com','印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()

            if status == '2':
                if orm.usage == u'使用':
                    if add_process:
                        orm.status = 3
                        orm.approve_now = add_process
                        email_orm = user_table.objects.get(name=add_process)
                        Thread(target=send_mail,args=(email_orm.email,'印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                    else:
                        orm.status = 5
                        orm.approve_now = ''
                else:
                    if add_process:
                        orm.status = 3
                        orm.approve_now = add_process
                        email_orm = user_table.objects.get(name=add_process)
                        Thread(target=send_mail,args=(email_orm.email,'印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                    else:
                        orm.status = 4

            if status == '3':
                if orm.usage == u'使用':
                    orm.status = 5
                    orm.approve_now = ''
                else:
                    orm.status = 4
                    if orm.seal_class == u'法人章':
                        orm.approve_now = u'龚晓芸'
                        Thread(target=send_mail,args=('gongxiaoyun@xiaohulu.com','印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
                    else:
                        orm.approve_now = u'吴佳伟'
                        Thread(target=send_mail,args=('zhangliying@xiaohulu.com','印章审核提醒','<h3>有一个印章审批等待您的处理，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000/KPI_table_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
            if status == '4':
                orm.status = 5
                orm.approve_now = ''


            # if status == '8':
            #     orm.status = 80
            #     orm.approve_now = orm.name
            #     orm.stamp_status = 1
            #
            # if status == '80':
            #     orm.status = 9
            #     orm.approve_now = status_owner[orm.status]
            #     orm.stamp_status = 1
            #
            # if status == '9':
            #     orm.status = 10
            #     orm.approve_now = ''
            #     orm.archive_status = 1

        if flag == '2':
            orm.status = -1
            orm.approve_now = orm.name


        if flag == '0':
            orm.status = 6
            orm.approve_now = ''

        orm.apply_time = datetime.datetime.now()
        orm.save()

        orm2 = detail(name=request.user.first_name,operation=flag,comment=comment,parent_id=_id)
        orm2.save()

        return HttpResponse(json.dumps({'code':0,'msg':u'审批成功'}),content_type="application/json")
    except Exception,e:
        import traceback
        print traceback.format_exc()
        print e




@login_required
def seal_process_detail_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = ['apply_time','name','operation','comment','id']

    try:
        parent_id = request.session['seal_id']
    except KeyError:
        parent_id = None
    if  sSortDir_0 == 'asc':
        if sSearch == '':
            result_data = detail.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).count()
        else:
            result_data = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)) \
                                            .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = detail.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).count()
        else:
            result_data = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)) \
                                            .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).filter(Q(name__contains=sSearch) | \
                                                                                          Q(comment__contains=sSearch) | \
                                                                                          Q(id__contains=sSearch)).count()

    for i in  result_data:
        aaData.append({
                       '0':str(i.apply_time),
                       '1':i.name,
                       '2':i.operation,
                       '3':i.comment.replace('\n','</br>'),
                       '4':i.id
                      })
    if parent_id:
        orm = table.objects.get(id=parent_id)
        if orm.status == 5 and orm.usage == u'外借':
            aaData[0]['2'] = 10
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")




@login_required
def seal_create_excel(request):
    _id = request.POST.get('id')
    orm = table.objects.get(id=_id)

    if orm.status != 5:
        return HttpResponse(json.dumps({'code':1,'msg':'您的流程未完成无法生成Excel'}),content_type="application/json")

    name = orm.name
    uuid = orm.uuid
    department = orm.department
    usage = orm.usage
    seal_from = orm.seal_from
    seal_class = orm.seal_class
    reason = orm.reason
    archive_path = orm.archive_path
    borrow_begin_time = orm.borrow_begin_time
    borrow_end_time = orm.borrow_end_time
    seal_time = str(borrow_begin_time) + '->' + str(borrow_end_time)
    comment = orm.comment

    orm2 = detail.objects.filter(parent_id=_id)
    process_detail_list = []
    for i in orm2:
        process_detail_list.append({'apply_time':i.apply_time,
                                    'name':i.name,
                                    'operation':i.operation,
                                    'comment':i.comment})

    try:
        wb = load_workbook(filename = BASE_DIR + '/static/files/seal_template.xlsx')
        ws = wb.active
        ws['B2'] = name
        ws['D2'] = uuid
        ws['B3'] = department
        ws['D3'] = usage
        ws['B4'] = seal_from
        ws['D4'] = seal_class
        ws['B5'] = reason
        ws['B6'] = archive_path
        ws['B7'] = seal_time
        ws['B8'] = comment

        row_num = 11
        for i in process_detail_list:
            ws['B'+str(row_num)] = i['apply_time']
            ws['C'+str(row_num)] = i['name']
            if i['operation'] == 1:
                operation = u'审批通过'
            elif i['operation'] == 0:
                operation = u'审批不通过'
            elif i['operation'] == -1:
                operation = u'退回修改'
            elif i['operation'] == 9:
                operation = u'印章申请'
            ws['D'+str(row_num)] = operation
            ws['E'+str(row_num)] = i['comment']
            row_num += 1

        wb.save(BASE_DIR + '/static/files/seal.xlsx')
        return HttpResponse(json.dumps({'code':0,'msg':u'生成成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")



@login_required
def seal_approve_alert(request):
    orm = table.objects.filter(approve_now=request.user.first_name)
    if len(orm) > 0:
        msg = '有%s个印章事件等待您的审批' % len(orm)
        return HttpResponse(json.dumps({'code':0,'msg':msg}),content_type="application/json")
    else:
        return HttpResponse(json.dumps({'code':1}),content_type="application/json")



@login_required
def export_seal_list(request):
    year = request.POST.get('year')
    seal_class = request.POST.get('seal_class')
    flag = ''
    if year:
        flag += 'x'
    if seal_class:
        flag += 'y'
    if flag:
        if flag == 'x':
            orm = table.objects.filter(apply_time__year=int(year)).filter(status=10)
        if flag == 'y':
            orm = table.objects.filter(seal_class=seal_class).filter(status=10)
        if flag == 'xy':
            orm = table.objects.filter(apply_time__year=int(year)).filter(seal_class=seal_class).filter(status=10)
    else:
        orm = table.objects.filter(status=5)
    contract_list = []
    for i in orm:
        contract_list.append({'uuid':i.uuid,
                              'name':i.name,
                              'department':i.department,
                              'seal_class':i.seal_class,
                              'usage':i.usage})
    try:
        wb = load_workbook(filename = BASE_DIR + '/static/files/seal_list_template.xlsx')
        ws = wb.active

        row_num = 3
        for i in contract_list:
            ws['A'+str(row_num)] = i['uuid']
            ws['B'+str(row_num)] = i['name']
            ws['C'+str(row_num)] = i['department']
            ws['D'+str(row_num)] = i['seal_class']
            ws['E'+str(row_num)] = i['usage']
            row_num += 1

        wb.save(BASE_DIR + '/static/files/seal_list.xlsx')
        return HttpResponse(json.dumps({'code':0,'msg':u'生成成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")