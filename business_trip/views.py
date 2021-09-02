# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.http import HttpResponse,HttpResponseRedirect
from django.template import RequestContext
from django.db.models.query_utils import Q
from django.contrib.auth.decorators import login_required
from libs.sendmail import send_mail
from business_trip.models import *
from vacation.models import user_table
import json
import datetime
from wzhl_oa.settings import administration, BASE_DIR
import os
from threading import Thread
from openpyxl import load_workbook


import sys
reload(sys)
sys.setdefaultencoding('utf-8')


@login_required
def business_trip_table(request):
    path = request.path.split('/')[1]

    # orm = table.objects.filter(name=request.user.first_name)
    # for i in orm:
    #     if (datetime.datetime.now() - i.commit_time).days >= 20 and i.status not in (5, 6, 7, 8):
    #         i.status = 8
    #         i.save()

    return render(request, 'business_trip/business_trip_table.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'business_trip',
                                                 'path2':path,
                                                 'page_name1':u'出差管理',
                                                 'page_name2':u'出差申请',
                                                 'username':request.user.username},
                                                context_instance=RequestContext(request))



@login_required
def business_trip_table_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = [None, None, None, 'apply_time', None, None, 'budget_sum', 'status', None, None, 'id', None]

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            print iSortCol_0
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name).filter(Q(name__contains=sSearch)) \
                                                                                .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).filter(Q(name__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = table.objects.filter(name=request.user.first_name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).count()
        else:
            result_data = table.objects.filter(name=request.user.first_name).filter(Q(name__contains=sSearch)) \
                                                    .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = table.objects.filter(name=request.user.first_name).filter(Q(name__contains=sSearch)).count()



    for i in  result_data:
        budget_info = '''
                    <a class="btn btn-sm blue">
                        详细预算 <i class="fa"></i>
                    </a>
                '''
        export = '''
                    <a class="btn btn-sm green">
                        生成Excel文件 <i class="fa fa-level-down"></i>
                    </a>
                '''
        if i.status != 0:
            delete = '''
                        <a class="btn btn-sm red">
                            删除 <i class="fa"></i>
                        </a>
                     '''
        else:
            delete = '''
                        <a class="btn btn-sm green">
                            提交 <i class="fa"></i>
                        </a>
                     '''
        detail = '''
                    <a class="btn btn-sm blue">
                        行程详情 <i class="fa"></i>
                    </a>
                 '''
        # if i.hotel_reservation == 1:
        #     hotel_reservation = u'是'
        # else:
        #     hotel_reservation = u'否'
        aaData.append({
                       '0':i.name,
                       '1':detail,
                       '2':budget_info,
                       '3':i.date,
                       '4':str(i.apply_time),
                       '5':i.travel_partner,
                       '6':i.budget_sum,
                       '7':i.status,
                       '8':export,
                       '9':delete,
                       '10':i.id,
                       '11':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")



@login_required
def business_trip_table_save(request):
    # reason = request.POST.get('reason')
    # destination = request.POST.get('destination')
    # begin = request.POST.get('begin')
    # end = request.POST.get('end')
    travel_partner = request.POST.get('travel_partner')
    # vehicle = request.POST.get('vehicle')
    hotel_reservation = request.POST.get('hotel_reservation')

    orm = table(name=request.user.first_name, date='', travel_partner=travel_partner,
         hotel_reservation=hotel_reservation, budget_sum=0, status=0, approve_now='',
         commit_time=datetime.datetime.now())
    orm.save()

    return HttpResponse(json.dumps({'code':0,'msg':u'保存成功'}),content_type="application/json")



@login_required
def business_trip_table_del(request):
    _id = request.POST.get('id')
    orm = table.objects.get(id=_id)
    if orm.status == 5 or orm.status == 6 or orm.status == 7 or orm.status == 8:
        return HttpResponse(json.dumps({'code':1,'msg':u'无法删除'}),content_type="application/json")
    else:
        orm.delete()
        return HttpResponse(json.dumps({'code':0,'msg':u'删除成功'}),content_type="application/json")





@login_required
def business_trip_set_session(request):
    parent_id = request.POST.get('parent_id')
    status = request.POST.get('status')
    request.session['parent_id'] = parent_id
    request.session['status'] = status
    return HttpResponse(json.dumps('OK'),content_type="application/json")





@login_required
def business_trip_budget(request):
    path = request.path.split('/')[1]
    return render(request, 'business_trip/business_trip_budget.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'business_trip',
                                                 'path2':path,
                                                 'page_name1':u'出差管理',
                                                 'page_name2':u'出差申请',
                                                 'username':request.user.username},
                                                context_instance=RequestContext(request))




@login_required
def business_trip_budget_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    parent_id = request.session.get('parent_id')

    aaData = []
    sort = ['budget_type', 'budget', 'id']

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            print iSortCol_0
            result_data = budget.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget.objects.filter(parent_id=parent_id).count()
        else:
            result_data = budget.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)) \
                                                                                .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = budget.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget.objects.filter(parent_id=parent_id).count()
        else:
            result_data = budget.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)) \
                                                    .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)).count()



    for i in  result_data:
        aaData.append({
                       '0':i.budget_type,
                       '1':i.budget,
                       '2':i.id,
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")



@login_required
def business_trip_budget_save(request):
    budget_type = request.POST.get('budget_type')
    _budget = request.POST.get('budget')
    parent_id = request.POST.get('parent_id')
    _id = request.POST.get('id')

    try:
        if not _id:
            orm = budget.objects.filter(parent_id=parent_id)
            for i in orm:
                if i.budget_type == budget_type:
                    return HttpResponse(json.dumps({'code':1,'msg':u'类型已存在'}),content_type="application/json")
            orm2 = budget(budget_type=budget_type, budget=_budget, parent_id=int(parent_id))
            orm2.save()
            budget_sum = 0
            orm3 = budget.objects.filter(parent_id=parent_id)
            for i in orm3:
                budget_sum += i.budget
            orm4 = table.objects.get(id=parent_id)
            orm4.budget_sum = budget_sum
            orm4.save()
        else:
            orm = budget.objects.get(id=_id)
            orm.budget_type = budget_type
            orm.budget = int(_budget)
            orm.save()
            budget_sum = 0
            orm2 = budget.objects.filter(parent_id=parent_id)
            for i in orm2:
                budget_sum += i.budget
            orm3 = table.objects.get(id=parent_id)
            orm3.budget_sum = budget_sum
            orm3.save()

        return HttpResponse(json.dumps({'code':0,'msg':u'保存成功'}),content_type="application/json")
    except Exception,e:
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")




@login_required
def business_trip_budget_del(request):
    _id = request.POST.get('id')
    orm = budget.objects.get(id=_id)
    orm.delete()
    return HttpResponse(json.dumps({'code':0,'msg':u'删除成功'}),content_type="application/json")





@login_required
def business_trip_budget_sub_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    parent_id = request.session.get('parent_id')

    aaData = []
    sort = ['airfare', 'accommodation_cost', 'payment', 'id']

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            print iSortCol_0
            result_data = budget_sub.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget_sub.objects.filter(parent_id=parent_id).count()
        else:
            result_data = budget_sub.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)) \
                                                                                .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget_sub.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = budget_sub.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget_sub.objects.filter(parent_id=parent_id).count()
        else:
            result_data = budget_sub.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)) \
                                                    .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = budget_sub.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)).count()



    for i in  result_data:
        aaData.append({
                       '0':i.airfare,
                       '1':i.accommodation_cost,
                       '2':i.payment,
                       '3':i.id,
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")



@login_required
def business_trip_budget_sub_save(request):
    airfare = request.POST.get('airfare')
    accommodation_cost = request.POST.get('accommodation_cost')
    payment = request.POST.get('payment')
    parent_id = request.POST.get('parent_id')
    _id = request.POST.get('id')

    try:
        if not _id:
            orm = budget_sub(airfare=airfare, accommodation_cost=accommodation_cost, payment=payment, parent_id=int(parent_id))
            orm.save()
        else:
            orm = budget_sub.objects.get(id=_id)
            orm.airfare = airfare
            orm.accommodation_cost = accommodation_cost
            orm.payment = payment
            orm.save()

        return HttpResponse(json.dumps({'code':0,'msg':u'保存成功'}),content_type="application/json")
    except Exception,e:
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")




@login_required
def business_trip_budget_sub_del(request):
    _id = request.POST.get('id')
    orm = budget_sub.objects.get(id=_id)
    orm.delete()
    return HttpResponse(json.dumps({'code':0,'msg':u'删除成功'}),content_type="application/json")






@login_required
def business_trip_detail(request):
    path = request.path.split('/')[1]
    return render(request, 'business_trip/business_trip_detail.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'business_trip',
                                                 'path2':path,
                                                 'page_name1':u'出差管理',
                                                 'page_name2':u'出差申请',
                                                 'username':request.user.username},
                                                context_instance=RequestContext(request))




@login_required
def business_trip_detail_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    parent_id = request.session.get('parent_id')

    aaData = []
    sort = ['reason', 'origin', 'destination', None, 'vehicle', 'id']

    if  sSortDir_0 == 'asc':
        if sSearch == '':
            print iSortCol_0
            result_data = detail.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).count()
        else:
            result_data = detail.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)) \
                                                                                .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)).count()
    else:
        if sSearch == '':
            result_data = detail.objects.filter(parent_id=parent_id).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).count()
        else:
            result_data = detail.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)) \
                                                    .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
            iTotalRecords = detail.objects.filter(parent_id=parent_id).filter(Q(budget_type__contains=sSearch) | \
                                                                                    Q(budget__contains=sSearch)).count()



    for i in  result_data:
        aaData.append({
                       '0':i.reason,
                       '1':i.origin,
                       '2':i.destination,
                       '3':i.date,
                       '4':i.vehicle,
                       '5':i.id,
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")



@login_required
def business_trip_detail_save(request):
    reason = request.POST.get('reason')
    origin = request.POST.get('origin')
    destination = request.POST.get('destination')
    begin = request.POST.get('begin')
    end = request.POST.get('end')
    vehicle = request.POST.get('vehicle')
    parent_id = request.POST.get('parent_id')
    _id = request.POST.get('id')

    if begin == end:
        date = begin
    else:
        date = begin + '&nbsp->&nbsp' + end

    if not _id:
        orm = detail(reason=reason, origin=origin, destination=destination, date=date, vehicle=vehicle, parent_id=parent_id)
        orm.save()
    else:
        orm = detail.objects.get(id=_id)
        orm.reason = reason
        orm.origin = origin
        orm.destination = destination
        orm.date = date
        orm.vehicle = vehicle
        orm.save()
    orm2 = detail.objects.filter(parent_id=parent_id)
    date_list = []
    for i in orm2:
        begin_date_list = i.date.split('&nbsp')[0].split('-')
        end_date_list = i.date.split('&nbsp')[-1].split('-')
        begin_datetime = datetime.date(int(begin_date_list[0]), int(begin_date_list[1]), int(begin_date_list[2]))
        end_datetime = datetime.date(int(end_date_list[0]), int(end_date_list[1]), int(end_date_list[2]))
        date_list.append(begin_datetime)
        date_list.append(end_datetime)
    total_date = str(min(date_list)) + '&nbsp->&nbsp' + str(max(date_list))
    orm3 = table.objects.get(id=parent_id)
    orm3.date = total_date
    orm3.save()
    return HttpResponse(json.dumps({'code':0,'msg':u'保存成功'}),content_type="application/json")




@login_required
def business_trip_detail_del(request):
    _id = request.POST.get('id')
    parent_id = request.POST.get('parent_id')

    orm = detail.objects.get(id=_id)
    orm.delete()
    orm2 = detail.objects.filter(parent_id=parent_id)
    date_list = []
    for i in orm2:
        begin_date_list = i.date.split('&nbsp')[0].split('-')
        end_date_list = i.date.split('&nbsp')[-1].split('-')
        begin_datetime = datetime.date(int(begin_date_list[0]), int(begin_date_list[1]), int(begin_date_list[2]))
        end_datetime = datetime.date(int(end_date_list[0]), int(end_date_list[1]), int(end_date_list[2]))
        date_list.append(begin_datetime)
        date_list.append(end_datetime)
    total_date = str(min(date_list)) + '&nbsp->&nbsp' + str(max(date_list))
    orm3 = table.objects.get(id=parent_id)
    orm3.date = total_date
    orm3.save()
    return HttpResponse(json.dumps({'code':0,'msg':u'删除成功'}),content_type="application/json")




@login_required
def business_trip_approve(request):
    path = request.path.split('/')[1]
    return render(request, 'business_trip/business_trip_approve.html',{'user':'%s%s' % (request.user.last_name,request.user.first_name),
                                                 'path1':'business_trip',
                                                 'path2':path,
                                                 'page_name1':u'出差管理',
                                                 'page_name2':u'出差审批',
                                                 'username':request.user.username},
                                                context_instance=RequestContext(request))





@login_required
def business_trip_approve_data(request):
    sEcho =  request.POST.get('sEcho') #标志，直接返回
    iDisplayStart = int(request.POST.get('iDisplayStart'))#第几行开始
    iDisplayLength = int(request.POST.get('iDisplayLength'))#显示多少行
    iSortCol_0 = int(request.POST.get("iSortCol_0"))#排序行号
    sSortDir_0 = request.POST.get('sSortDir_0')#asc/desc
    sSearch = request.POST.get('sSearch')#高级搜索

    aaData = []
    sort = [None, None, None, 'apply_time', None, None, None, None, 'status', None, None, None, 'id', None]

    if request.user.has_perm('business_trip.can_view_all'):
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                print iSortCol_0
                result_data = table.objects.all().order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all().filter(Q(name__contains=sSearch)) \
                                                                                    .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().filter(Q(name__contains=sSearch)).count()
        else:
            if sSearch == '':
                result_data = table.objects.all().order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().count()
            else:
                result_data = table.objects.all().filter(Q(name__contains=sSearch)) \
                                                        .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.all().filter(Q(name__contains=sSearch)).count()
    else:
        if  sSortDir_0 == 'asc':
            if sSearch == '':
                print iSortCol_0
                result_data = table.objects.filter(approve_now=request.user.first_name).order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(approve_now=request.user.first_name).count()
            else:
                result_data = table.objects.filter(approve_now=request.user.first_name).filter(Q(name__contains=sSearch)) \
                                                                                    .order_by(sort[iSortCol_0])[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(approve_now=request.user.first_name).filter(Q(name__contains=sSearch)).count()
        else:
            if sSearch == '':
                result_data = table.objects.filter(approve_now=request.user.first_name).order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(approve_now=request.user.first_name).count()
            else:
                result_data = table.objects.filter(approve_now=request.user.first_name).filter(Q(name__contains=sSearch)) \
                                                        .order_by(sort[iSortCol_0]).reverse()[iDisplayStart:iDisplayStart+iDisplayLength]
                iTotalRecords = table.objects.filter(approve_now=request.user.first_name).filter(Q(name__contains=sSearch)).count()


    for i in  result_data:
        detail = '''
                    <a class="btn btn-sm blue">
                        行程详情 <i class="fa"></i>
                    </a>
                 '''
        budget_info = '''
                    <a class="btn btn-sm blue">
                        详细预算 <i class="fa"></i>
                    </a>
                '''
        export = '''
                    <a class="btn btn-sm green">
                        生成Excel文件 <i class="fa fa-level-down"></i>
                    </a>
                '''
        approve = '''
                    <a class="btn btn-sm green">
                        审批通过 <i class="fa"></i>
                    </a>
                 '''
        not_approve = '''
                    <a class="btn btn-sm red">
                        审批不通过 <i class="fa"></i>
                    </a>
                 '''
        if i.hotel_reservation == 1:
            hotel_reservation = u'是'
        else:
            hotel_reservation = u'否'
        aaData.append({
                       '0':i.name,
                       '1':detail,
                       '2':budget_info,
                       '3':i.date,
                       '4':str(i.apply_time),
                       '5':i.travel_partner,
                       '6':hotel_reservation,
                       '7':i.budget_sum,
                       '8':i.status,
                       '9':export,
                       '10':approve,
                       '11':not_approve,
                       '12':i.id,
                       '13':i.approve_now
                      })
    result = {'sEcho':sEcho,
               'iTotalRecords':iTotalRecords,
               'iTotalDisplayRecords':iTotalRecords,
               'aaData':aaData
    }
    return HttpResponse(json.dumps(result),content_type="application/json")



@login_required
def business_trip_approve_process(request):
    flag = request.POST.get('flag')
    _id = request.POST.get('id')
    status = request.POST.get('status')

    orm = table.objects.get(id=_id)
    if flag == '1':
        if status == '0':
            if orm.budget_sum == 0:
                return HttpResponse(json.dumps({'code':1,'msg':u'请填写详细预算'}),content_type="application/json")
            orm2 = detail.objects.filter(parent_id=_id)
            if len(orm2) <= 0:
                return HttpResponse(json.dumps({'code':1,'msg':u'请填写行程详情'}),content_type="application/json")
            orm3 = user_table.objects.get(name=request.user.first_name)
            if orm3.supervisor == orm3.principal:
                orm.status = 2
            else:
                orm.status = 1
            orm.approve_now = orm3.supervisor
            orm.apply_time = datetime.datetime.now()

            orm4 = user_table.objects.get(name=orm3.supervisor)
            email = orm4.email
            Thread(target=send_mail,args=(email,'出差审批提醒','<h3>有一个出差事件等待您的审批，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000:10000/business_trip_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
        elif status == '1':
            orm.status = 2
            orm2 = user_table.objects.get(name=request.user.first_name)
            orm.approve_now = orm2.principal
            orm.apply_time = datetime.datetime.now()

            orm3 = user_table.objects.get(name=orm2.principal)
            email = orm3.email
            Thread(target=send_mail,args=(email,'出差审批提醒','<h3>有一个出差事件等待您的审批，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000:10000/business_trip_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
        elif status == '2':
            orm.status = 3
            orm.approve_now = administration['name']
            orm.apply_time = datetime.datetime.now()

            Thread(target=send_mail,args=(administration['email'],'出差审批提醒','<h3>有一个出差事件等待您的审批，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000:10000/business_trip_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
        elif status == '3':
            orm.status = 4
            orm.approve_now = administration['name']
            orm.apply_time = datetime.datetime.now()

            Thread(target=send_mail,args=(administration['email'],'出差审批提醒','<h3>有一个出差事件等待您的审批，请在OA系统中查看。</h3><br>OA链接：http://oa.xiaoquan.com:10000:10000/business_trip_approve/</br><br>此邮件为自动发送的提醒邮件，请勿回复。')).start()
        elif status == '4':
            orm2 = budget_sub.objects.filter(parent_id=_id)
            if len(orm2) <= 0:
                return HttpResponse(json.dumps({'code':1,'msg':u'请填实际预算'}),content_type="application/json")
            orm.status = 5
            orm.approve_now = ''
    else:
        if status == '4':
            orm.status = 6
            orm.approve_now = ''
        else:
            orm.status = 0
            orm.approve_now = ''
            orm.apply_time = datetime.datetime.now()

    orm.save()
    return HttpResponse(json.dumps({'code':0,'msg':u'提交成功'}),content_type="application/json")





@login_required
def business_trip_create_excel(requests):
    _id = requests.POST.get('id')
    orm = table.objects.get(id=_id)

    if orm.status != 5:
        return HttpResponse(json.dumps({'code':1,'msg':'您的流程未完成无法生成Excel'}),content_type="application/json")

    name = orm.name
    apply_date = str(orm.apply_time)
    orm2 = user_table.objects.get(name=name)
    department = orm2.department
    supervisor = orm2.supervisor
    principal = orm2.principal
    travel_partner = orm.travel_partner
    name_apply_date = name + ' ' + apply_date

    detail_list = []
    orm3 = detail.objects.filter(parent_id=_id).order_by('id')
    for i in orm3:
        detail_list.append({'date': i.date.replace('&nbsp',' '),
                            'origin': i.origin,
                            'destination': i.destination,
                            'vehicle': i.vehicle,
                            'reason': i.reason})

    budget_dict = {'1': 0,
                   '2': 0,
                   '3': 0,
                   '4': 0}
    budget_sum = 0
    orm4 = budget.objects.filter(parent_id=_id)
    for i in orm4:
        if i.budget_type == '交通费':
            budget_dict['1'] = i.budget
        elif i.budget_type == '膳食费':
            budget_dict['2'] = i.budget
        elif i.budget_type == '招待费':
            budget_dict['3'] = i.budget
        elif i.budget_type == '其他费用':
            budget_dict['4'] = i.budget
        budget_sum += i.budget


    orm5 = budget_sub.objects.get(parent_id=_id)
    airfare = orm5.airfare
    accommodation_cost = orm5.accommodation_cost
    payment = orm5.payment

    try:
        wb = load_workbook(filename = BASE_DIR + '/static/files/business_trip_template.xlsx')
        ws = wb.active
        ws['A7'] = department
        ws['L7'] = name
        ws['W7'] = travel_partner

        row_num = '11'
        for i in detail_list:
            ws['D'+row_num] = i['date']
            ws['L'+row_num] = i['origin']
            ws['P'+row_num] = i['destination']
            ws['T'+row_num] = i['vehicle']
            ws['W'+row_num] = i['reason']
            row_num = str(int(row_num) + 1)

        ws['A18'] = budget_dict['1']
        ws['G18'] = budget_dict['2']
        ws['N18'] = budget_dict['3']
        ws['T18'] = budget_dict['4']
        ws['AA18'] = budget_sum

        ws['A21'] = airfare
        ws['L21'] = accommodation_cost
        ws['W21'] = payment

        ws['E25'] = name_apply_date
        ws['Q25'] = supervisor
        ws['AB25'] = principal

        wb.save(BASE_DIR + '/static/files/business_trip.xlsx')
        return HttpResponse(json.dumps({'code':0,'msg':u'生成成功'}),content_type="application/json")
    except Exception,e:
        print e
        return HttpResponse(json.dumps({'code':1,'msg':str(e)}),content_type="application/json")



@login_required
def business_trip_approve_alert(request):
    orm = table.objects.filter(approve_now=request.user.first_name)
    if len(orm) > 0:
        msg = '有%s个出差事件等待您的审批' % len(orm)
        return HttpResponse(json.dumps({'code':0,'msg':msg}),content_type="application/json")
    else:
        return HttpResponse(json.dumps({'code':1}),content_type="application/json")